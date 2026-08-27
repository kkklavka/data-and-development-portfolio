"""
Приложение для подготовки отчёта «Обнуление» из Excel-выгрузки 1С. Версия 19.

Что делает:
1. Открывает один или несколько Excel-файлов формата «Ведомость по товарам на складах».
2. Если колонок «Темп продаж», «Остаток товара в днях» и «Рекомендуемый заказ» нет — добавляет их.
3. Считает:
   - Темп продаж = Расход / количество дней в периоде отчёта;
   - Остаток товара в днях = Конечный остаток / Темп продаж;
   - Рекомендуемый заказ = максимум 0 или Темп продаж × дни до следующего завоза - Конечный остаток.
   - К ячейке «Рекомендуемый заказ» добавляет примечание: берём/не берём/проверить и причину.
     Примечания добавляются после сортировки, чтобы они не съезжали на другие товары.
4. Подсвечивает ТОЛЬКО два случая:
   - «Конечный остаток» красным, если он равен 0;
   - «Остаток товара в днях» красным, если он меньше 7.
5. Сортирует товары по колонке «Остаток товара в днях» по возрастанию, чтобы критичные были сверху.
6. Ставит фильтр как в готовом примере: строка 10, товары уже отсортированы, служебные строки 1–8 скрыты.
7. Сохраняет готовые Excel-файлы в выбранную папку.
8. Пустой конечный остаток считает как 0, ставит 0 в ячейку и подсвечивает.
9. Обрабатывает файлы в отдельном потоке, чтобы окно не зависало с надписью «Не отвечает».

Перед запуском один раз установи библиотеку:
pip install openpyxl pywin32
"""

import os
import sys
import re
import copy
import math
from datetime import datetime
import zipfile
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
import threading

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
NO_FILL = PatternFill(fill_type=None)

DEFAULT_SUPPLIER_FILENAMES = ["supplier_map.xlsx", "справочник товар - поставщик.xlsx"]
DEFAULT_SCHEDULE_FILENAMES = ["delivery_schedule.xlsx", "график поставок.xlsx"]


def app_dir():
    """Папка приложения.

    В обычном запуске это папка с .py-файлом.
    В собранном .exe через PyInstaller это временная папка распаковки,
    где лежат встроенные supplier_map.xlsx и delivery_schedule.xlsx.
    """
    try:
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            return sys._MEIPASS
        return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        return os.getcwd()


def default_file_path(filenames):
    """Ищет файл рядом с программой. Сначала латинское имя, потом старое русское имя.
    Латинские имена нужны, чтобы Windows/архиватор не ломал кириллицу при распаковке.
    """
    if isinstance(filenames, str):
        filenames = [filenames]
    for filename in filenames:
        path = os.path.join(app_dir(), filename)
        if os.path.exists(path):
            return path
    return ""



def make_xlsx_openpyxl_compatible(input_path):
    """
    Исправляет типичную проблему выгрузок 1С/Excel 2007:
    внутри .xlsx файл sharedStrings может лежать с другим регистром имени,
    например xl/SharedStrings.xml, а openpyxl ищет строго xl/sharedStrings.xml.

    Исходный файл НЕ меняется: создаётся временная копия.
    """
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp.name
    temp.close()

    with zipfile.ZipFile(input_path, 'r') as zin:
        names = zin.namelist()
        exact_shared = 'xl/sharedStrings.xml'

        real_shared_name = None
        for name in names:
            if name.lower() == 'xl/sharedstrings.xml':
                real_shared_name = name
                break

        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            wrote_shared = False
            for item in zin.infolist():
                original_name = item.filename
                data = zin.read(original_name)
                target_name = original_name

                # В ZIP имена чувствительны к регистру. Перезаписываем путь sharedStrings строго как ждёт openpyxl.
                if original_name.lower() == 'xl/sharedstrings.xml':
                    target_name = exact_shared
                    wrote_shared = True

                if original_name == '[Content_Types].xml':
                    text = data.decode('utf-8', errors='ignore')
                    text = text.replace('/xl/SharedStrings.xml', '/xl/sharedStrings.xml')
                    text = text.replace('/xl/sharedstrings.xml', '/xl/sharedStrings.xml')
                    if real_shared_name and 'sharedStrings+xml' not in text:
                        override = '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                        text = text.replace('</Types>', override + '</Types>')
                    data = text.encode('utf-8')

                # Если в архиве уже был путь с неправильным регистром, не дублируем его вторым файлом.
                zout.writestr(target_name, data)

            if real_shared_name and not wrote_shared:
                zout.writestr(exact_shared, zin.read(real_shared_name))

            if not real_shared_name and exact_shared not in names:
                empty_shared_strings = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"></sst>'
                )
                zout.writestr(exact_shared, empty_shared_strings)

    return temp_path

def resave_with_installed_excel(input_path):
    """
    Для самых кривых выгрузок 1С: открывает файл настоящим Excel и сохраняет
    временную нормальную копию .xlsx, которую потом уже читает openpyxl.

    Работает только на Windows, если установлен Microsoft Excel.
    Нужна библиотека pywin32:
    python -m pip install pywin32
    """
    try:
        import pythoncom
        import win32com.client
    except Exception as e:
        raise ParseError(
            "Файл открывается в Excel, но Python не может прочитать его напрямую. "
            "Установи дополнительную библиотеку: python -m pip install pywin32. "
            "После этого запусти программу ещё раз."
        ) from e

    input_path = os.path.abspath(input_path)
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = os.path.abspath(temp.name)
    temp.close()
    try:
        os.remove(temp_path)
    except Exception:
        pass

    excel = None
    wb = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        wb = excel.Workbooks.Open(input_path, UpdateLinks=0, ReadOnly=True, CorruptLoad=1)
        # 51 = xlOpenXMLWorkbook, обычная Книга Excel (*.xlsx)
        wb.SaveAs(temp_path, FileFormat=51)
        wb.Close(SaveChanges=False)
        wb = None
        excel.Quit()
        excel = None
        return temp_path
    except Exception as e:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        raise ParseError(
            "Не удалось автоматически пересохранить файл через Excel. "
            "Открой его вручную в Excel и сделай: Файл → Сохранить как → Книга Excel (*.xlsx), "
            "потом выбери пересохранённую копию."
        ) from e
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

def load_workbook_safely(input_path):
    """Открывает Excel. Если выгрузка 1С кривая — пробует несколько способов исправления."""
    temp_paths = []
    try:
        return load_workbook(input_path)
    except Exception as first_error:
        # 1) Быстрое ZIP-исправление для sharedStrings.
        try:
            temp_path = make_xlsx_openpyxl_compatible(input_path)
            temp_paths.append(temp_path)
            return load_workbook(temp_path)
        except Exception as second_error:
            # 2) Самый надёжный вариант на Windows: попросить настоящий Excel пересохранить файл.
            try:
                excel_temp_path = resave_with_installed_excel(input_path)
                temp_paths.append(excel_temp_path)
                return load_workbook(excel_temp_path)
            except ParseError as third_error:
                raise third_error from second_error
            except Exception as third_error:
                if 'sharedStrings' in str(first_error) or 'sharedStrings' in str(second_error):
                    raise ParseError(
                        "Не удалось открыть выгрузку из-за служебных строк Excel. "
                        "Открой файл в Excel и пересохрани через: Файл → Сохранить как → Книга Excel (*.xlsx)."
                    ) from third_error
                raise first_error
    finally:
        for temp_path in temp_paths:
            try:
                os.remove(temp_path)
            except Exception:
                pass


class ParseError(Exception):
    pass


def to_number(value):
    """Аккуратно превращает значение из Excel в число."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.startswith("#"):
        return 0.0

    text = text.replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0



def norm_text(value):
    """Нормализация для состыковки товаров/поставщиков из разных выгрузок 1С."""
    if value is None:
        return ""
    text = str(value).strip().lower().replace("ё", "е")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    # Убираем кавычки/лишнюю пунктуацию по краям, но не трогаем смысловые цифры и буквы.
    text = text.strip(" .,:;\t\n\r")
    return text


def find_column_by_header(ws, header_words, max_rows=80):
    """Ищет колонку по набору слов в заголовке."""
    for row in range(1, min(ws.max_row, max_rows) + 1):
        for col in range(1, ws.max_column + 1):
            value = norm_text(ws.cell(row=row, column=col).value)
            if not value:
                continue
            ok = True
            for word in header_words:
                if norm_text(word) not in value:
                    ok = False
                    break
            if ok:
                return row, col
    return None, None


WEEKDAY_NUM = {
    "понедельник": 0,
    "вторник": 1,
    "среда": 2,
    "четверг": 3,
    "пятница": 4,
    "суббота": 5,
    "воскресенье": 6,
}

WEEKDAY_SHORT = {
    0: "пн",
    1: "вт",
    2: "ср",
    3: "чт",
    4: "пт",
    5: "сб",
    6: "вс",
}


def parse_supplier_product_map(path):
    """
    Читает общий файл, где есть связь: Номенклатура -> Основной поставщик.
    Поддерживает выгрузку 1С, где эти колонки могут быть в середине листа.
    """
    if not path:
        return {}

    wb = load_workbook_safely(path)
    ws = wb.active

    header_row = None
    nomen_col = None
    supplier_col = None

    for row in range(1, min(ws.max_row, 120) + 1):
        found_nomen = None
        found_supplier = None
        for col in range(1, ws.max_column + 1):
            text = norm_text(ws.cell(row=row, column=col).value)
            if text == "номенклатура":
                found_nomen = col
            if "основной поставщик" in text:
                found_supplier = col
        if found_nomen and found_supplier:
            header_row = row
            nomen_col = found_nomen
            supplier_col = found_supplier
            break

    if not header_row:
        raise ParseError("В файле поставщиков не нашла колонки 'Номенклатура' и 'Основной поставщик'.")

    result = {}
    for row in range(header_row + 1, ws.max_row + 1):
        product = ws.cell(row=row, column=nomen_col).value
        supplier = ws.cell(row=row, column=supplier_col).value
        product_norm = norm_text(product)
        supplier_norm = norm_text(supplier)

        if not product_norm or product_norm == "итог":
            continue
        if not supplier_norm:
            continue
        if product_norm in ("количество (в базовых единицах)", "начальный остаток", "расход", "конечный остаток"):
            continue
        if "количество" in product_norm and len(product_norm) < 40:
            continue

        # Если товар встречается несколько раз, оставляем первого найденного поставщика.
        result.setdefault(product_norm, str(supplier).strip())

    if not result:
        raise ParseError("В файле поставщиков не нашла строк с товарами и основными поставщиками.")

    return result


def parse_delivery_schedule(path):
    """
    Читает график поставок: День поставки / Завоз / Поставщик.
    Возвращает поставщик -> дни недели, когда он приезжает.
    """
    if not path:
        return {}

    wb = load_workbook_safely(path)
    ws = wb.active

    day_col = None
    supplier_col = None
    header_row = None

    for row in range(1, min(ws.max_row, 80) + 1):
        possible_day_col = None
        possible_supplier_col = None
        for col in range(1, ws.max_column + 1):
            text = norm_text(ws.cell(row=row, column=col).value)
            if text in ("день поставки", "день недели"):
                possible_day_col = col
            if text == "поставщик":
                possible_supplier_col = col
        if possible_day_col and possible_supplier_col:
            day_col = possible_day_col
            supplier_col = possible_supplier_col
            header_row = row
            break

    if not header_row:
        raise ParseError("В графике поставок не нашла колонки 'День поставки' и 'Поставщик'.")

    schedule = {}
    for row in range(header_row + 1, ws.max_row + 1):
        day_text = norm_text(ws.cell(row=row, column=day_col).value)
        supplier = ws.cell(row=row, column=supplier_col).value
        supplier_norm = norm_text(supplier)
        if not supplier_norm:
            continue
        if day_text not in WEEKDAY_NUM:
            continue
        schedule.setdefault(supplier_norm, set()).add(WEEKDAY_NUM[day_text])

    if not schedule:
        raise ParseError("В графике поставок не нашла строк с поставщиками и днями поставки.")

    return schedule


def max_gap_between_delivery_days(day_numbers):
    """Самый длинный промежуток между поставками по кругу недели."""
    days = sorted(set(day_numbers))
    if not days:
        return None
    if len(days) == 1:
        return 7
    gaps = []
    for i in range(len(days) - 1):
        gaps.append(days[i + 1] - days[i])
    gaps.append(days[0] + 7 - days[-1])
    return max(gaps)


def build_supplier_order_info(schedule, safety_days=0):
    """Поставщик -> информация для расчёта заказа и примечаний."""
    result = {}
    for supplier_norm, day_numbers in schedule.items():
        days = sorted(set(day_numbers))
        gap = max_gap_between_delivery_days(days)
        if gap is not None:
            order_days = max(1, int(gap) + int(safety_days))
            result[supplier_norm] = {
                "order_days": order_days,
                "gap": int(gap),
                "safety_days": int(safety_days),
                "delivery_count": len(days),
                "delivery_days_text": ", ".join(WEEKDAY_SHORT.get(day, str(day)) for day in days),
            }
    return result


def build_supplier_order_days(schedule, safety_days=0):
    """Совместимость со старой логикой: поставщик -> дни для заказа."""
    return {supplier: info["order_days"] for supplier, info in build_supplier_order_info(schedule, safety_days).items()}


RUSSIAN_MONTHS = {
    "января": 1, "январь": 1, "янв": 1,
    "февраля": 2, "февраль": 2, "фев": 2,
    "марта": 3, "март": 3, "мар": 3,
    "апреля": 4, "апрель": 4, "апр": 4,
    "мая": 5, "май": 5,
    "июня": 6, "июнь": 6, "июн": 6,
    "июля": 7, "июль": 7, "июл": 7,
    "августа": 8, "август": 8, "авг": 8,
    "сентября": 9, "сентябрь": 9, "сен": 9, "сент": 9,
    "октября": 10, "октябрь": 10, "окт": 10,
    "ноября": 11, "ноябрь": 11, "ноя": 11,
    "декабря": 12, "декабрь": 12, "дек": 12,
    # Казахские названия — на случай локализованной выгрузки 1С.
    "қаңтар": 1, "ақпан": 2, "наурыз": 3, "сәуір": 4,
    "мамыр": 5, "маусым": 6, "шілде": 7, "тамыз": 8,
    "қыркүйек": 9, "қазан": 10, "қараша": 11, "желтоқсан": 12,
}


def _dates_from_text(value):
    """Возвращает все даты, которые удалось распознать в строке 1С."""
    text = str(value).strip().lower().replace("ё", "е")
    result = []

    # Числовые форматы: 16.07.2026, 16/07/2026, 16-07-2026.
    for day, month, year in re.findall(r"(?<!\d)(\d{1,2})[./-](\d{1,2})[./-](\d{4})(?!\d)", text):
        try:
            result.append(datetime(int(year), int(month), int(day)).date())
        except ValueError:
            pass

    # Текстовые форматы: 16 июля 2026 г.; 16 шілде 2026 ж.
    month_names = "|".join(sorted((re.escape(name) for name in RUSSIAN_MONTHS), key=len, reverse=True))

    # Компактный диапазон одного месяца: 12–15 июля 2026 г. / с 12 по 15 июля 2026 г.
    compact_range = rf"(?<!\d)(\d{{1,2}})\s*(?:-|–|—|по)\s*(\d{{1,2}})\s+({month_names})\.?\s+(\d{{4}})(?!\d)"
    for start_day, end_day, month_name, year in re.findall(compact_range, text, flags=re.IGNORECASE):
        try:
            month = RUSSIAN_MONTHS[month_name.lower()]
            result.append(datetime(int(year), month, int(start_day)).date())
            result.append(datetime(int(year), month, int(end_day)).date())
        except (ValueError, KeyError):
            pass

    pattern = rf"(?<!\d)(\d{{1,2}})\s+({month_names})\.?\s+(\d{{4}})(?!\d)"
    for day, month_name, year in re.findall(pattern, text, flags=re.IGNORECASE):
        try:
            result.append(datetime(int(year), RUSSIAN_MONTHS[month_name.lower()], int(day)).date())
        except (ValueError, KeyError):
            pass

    # Убираем дубликаты, сохраняя порядок появления.
    unique = []
    for parsed_date in result:
        if parsed_date not in unique:
            unique.append(parsed_date)
    return unique


def _period_length(start, end):
    """Возвращает календарное число дней в периоде, включая обе граничные даты."""
    if end < start:
        start, end = end, start
    return (end - start).days + 1


def parse_period_days(ws):
    """Определяет продолжительность отчёта из разных форматов выгрузки 1С.

    Поддерживает, например:
    - «Период: 01.07.2026 - 07.07.2026»;
    - «Период: 16 июля 2026 г.»;
    - «Период: 16.07.2026»;
    - даты в строках заголовков, если отдельной строки «Период» нет.
    """
    max_row = min(ws.max_row, 25)
    max_col = min(ws.max_column, 30)

    # Сначала доверяем строке «Период», чтобы даты колонок не повлияли на расчёт.
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            if "период" not in str(value).lower():
                continue
            dates = _dates_from_text(value)
            if len(dates) >= 2:
                return _period_length(dates[0], dates[1])
            if len(dates) == 1:
                return 1

    # Запасной вариант: собираем даты из верхней части отчёта.
    found_dates = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            if hasattr(value, "date") and not isinstance(value, str):
                try:
                    parsed_date = value.date()
                    if parsed_date not in found_dates:
                        found_dates.append(parsed_date)
                except Exception:
                    pass
            for parsed_date in _dates_from_text(value):
                if parsed_date not in found_dates:
                    found_dates.append(parsed_date)

    if len(found_dates) >= 2:
        return _period_length(min(found_dates), max(found_dates))
    if len(found_dates) == 1:
        return 1

    raise ParseError(
        "Не удалось определить период отчёта. Поддерживаются диапазоны дат "
        "(например, 01.07.2026–07.07.2026) и отчёты за один день "
        "(например, 16 июля 2026 г.)."
    )


def find_header_row(ws):
    """Ищет строку заголовка, где в колонке B написано 'Номенклатура'."""
    for row in range(1, min(ws.max_row, 100) + 1):
        value = ws.cell(row=row, column=2).value
        if value and str(value).strip().lower() == "номенклатура":
            return row
    raise ParseError("Не нашла строку с заголовком 'Номенклатура'.")


def copy_style(source_cell, target_cell):
    """Копирует оформление ячейки."""
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy.copy(source_cell.protection)


def ensure_extra_columns(ws, header_row):
    """Добавляет/обновляет колонки G:I: Темп продаж, Остаток товара в днях, Рекомендуемый заказ."""
    subheader_row = header_row + 1

    # Если в файле только A:F, просто заполняем G:I. Если G:I уже есть — перезаписываем заголовки.
    ws.cell(row=subheader_row, column=7, value="Темп продаж")
    ws.cell(row=subheader_row, column=8, value="Остаток товара в днях")
    ws.cell(row=subheader_row, column=9, value="Рекомендуемый заказ")

    # Оформление новых заголовков берём с соседнего заголовка F.
    for col in (7, 8, 9):
        copy_style(ws.cell(row=subheader_row, column=6), ws.cell(row=subheader_row, column=col))
        ws.cell(row=subheader_row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Верхний объединённый заголовок «Количество...» расширяем до I, если он есть на C:F/G:H.
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == header_row and merged_range.max_row == header_row and merged_range.min_col == 3:
            ws.unmerge_cells(str(merged_range))
            break
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=9)
    ws.cell(row=header_row, column=3).value = "Количество (в базовых единицах)"
    ws.cell(row=header_row, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ширина колонок, чтобы всё читалось.
    ws.column_dimensions[get_column_letter(7)].width = 14
    ws.column_dimensions[get_column_letter(8)].width = 20
    ws.column_dimensions[get_column_letter(9)].width = 20


def find_data_rows(ws, header_row):
    """Возвращает номера строк с товарами, не включая пустую строку и 'Итог'."""
    data_start = header_row + 2
    rows = []

    for row in range(data_start, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        if name is None or str(name).strip() == "":
            continue
        if str(name).strip().lower() == "итог":
            break
        rows.append(row)

    if not rows:
        raise ParseError("Не нашла строки с товарами.")

    return rows


def clear_old_highlighting(ws, rows):
    """Убирает старую заливку только в рабочих колонках F:I, чтобы не было лишней подсветки."""
    for row in rows:
        for col in (6, 7, 8, 9):
            ws.cell(row=row, column=col).fill = copy.copy(NO_FILL)
            # цвет шрифта сбрасываем только в F и H, где могла быть красная подсветка
            if col in (6, 8):
                ws.cell(row=row, column=col).font = copy.copy(ws.cell(row=row, column=5).font)


def clear_existing_conditional_formatting(ws):
    """Удаляет старые правила условного форматирования, чтобы подсвечивалось только нужное."""
    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass


def calculate_and_highlight(ws, rows, period_days, fallback_order_days, product_supplier_map=None, supplier_order_info=None):
    zero_end_count = 0
    less_7_count = 0
    order_count = 0
    product_supplier_map = product_supplier_map or {}
    supplier_order_info = supplier_order_info or {}

    stats = {
        "supplier_found": 0,
        "schedule_found": 0,
        "fallback_used": 0,
        "product_not_found": 0,
        "supplier_without_schedule": 0,
    }

    for row in rows:
        product_name = ws.cell(row=row, column=2).value
        product_norm = norm_text(product_name)
        expense = to_number(ws.cell(row=row, column=5).value)       # Расход, E
        end_cell = ws.cell(row=row, column=6)
        end_original = end_cell.value
        end_was_blank = end_original is None or str(end_original).strip() == ""
        end_balance = to_number(end_original)   # Конечный остаток, F
        if end_was_blank:
            end_cell.value = 0

        daily_sales = expense / period_days if period_days > 0 else 0.0
        days_left = end_balance / daily_sales if daily_sales > 0 else None

        daily_sales_cell = ws.cell(row=row, column=7)
        days_left_cell = ws.cell(row=row, column=8)
        recommended_order_cell = ws.cell(row=row, column=9)

        # Дни для рекомендуемого заказа:
        # 1) если есть общий справочник товар -> поставщик и график поставок, считаем по поставщику;
        # 2) если не нашли товар/поставщика/график, берём запасное значение из окна.
        order_days = fallback_order_days
        supplier = None
        supplier_note = ""
        used_fallback_reason = ""
        if product_supplier_map and supplier_order_info:
            supplier = product_supplier_map.get(product_norm)
            if supplier:
                stats["supplier_found"] += 1
                supplier_norm = norm_text(supplier)
                info = supplier_order_info.get(supplier_norm)
                if info:
                    order_days = info["order_days"]
                    stats["schedule_found"] += 1
                    delivery_word = "раз" if info["delivery_count"] == 1 else "раза" if 2 <= info["delivery_count"] <= 4 else "раз"
                    supplier_note = (
                        f"Поставщик: {supplier}. Завоз: {info['delivery_count']} {delivery_word} в неделю "
                        f"({info['delivery_days_text']}). Самый длинный промежуток: {info['gap']} дн."
                    )
                    if info.get("safety_days", 0):
                        supplier_note += f" + страховой запас {info['safety_days']} дн."
                else:
                    stats["supplier_without_schedule"] += 1
                    stats["fallback_used"] += 1
                    used_fallback_reason = f"Для поставщика {supplier} не найден график поставок."
            else:
                stats["product_not_found"] += 1
                stats["fallback_used"] += 1
                used_fallback_reason = "Товар не найден в справочнике товар-поставщик."
        elif product_supplier_map or supplier_order_info:
            used_fallback_reason = "Не выбран один из общих файлов для расчёта по поставщикам."

        if daily_sales > 0:
            raw_order = daily_sales * order_days - end_balance
            recommended_order = math.ceil(max(0, raw_order))
        else:
            raw_order = 0.0
            recommended_order = 0

        daily_sales_cell.value = round(daily_sales, 4)
        days_left_cell.value = round(days_left, 2) if days_left is not None else ""
        recommended_order_cell.value = recommended_order

        daily_sales_cell.number_format = "0.00"
        days_left_cell.number_format = "0.00"
        recommended_order_cell.number_format = "0"

        # Чтобы новые значения выглядели как соседние числовые колонки.
        copy_style(ws.cell(row=row, column=6), daily_sales_cell)
        copy_style(ws.cell(row=row, column=6), days_left_cell)
        copy_style(ws.cell(row=row, column=6), recommended_order_cell)
        daily_sales_cell.number_format = "0.00"
        days_left_cell.number_format = "0.00"
        recommended_order_cell.number_format = "0"
        if recommended_order > 0:
            order_count += 1

        # Примечание к рекомендуемому заказу. Это не отдельная колонка — комментарий внутри ячейки.
        try:
            base_note = supplier_note if supplier_note else f"Расчёт по запасному значению: {order_days} дн."
            if used_fallback_reason:
                base_note += f" {used_fallback_reason}"

            if expense <= 0 or daily_sales <= 0:
                if end_balance > 0:
                    decision_note = "Не берём: за период не было продаж."
                elif end_balance == 0:
                    decision_note = "Не берём: товара нет на остатке, но за период не было продаж."
                else:
                    decision_note = "Проверить: отрицательный остаток, продаж за период не было."
                comment_text = (
                    f"{base_note}\n"
                    f"{decision_note}\n"
                    "Рекомендуемый заказ: 0, потому что темп продаж = 0."
                )
            elif recommended_order > 0:
                if end_balance <= 0:
                    decision_note = "Берём: продажи были, текущий остаток 0/минус. Расчёт до следующего завоза."
                else:
                    decision_note = "Берём: текущего остатка не хватит до следующего завоза."
                comment_text = (
                    f"{base_note}\n"
                    f"{decision_note}\n"
                    f"Расчёт: темп продаж {daily_sales:.2f} × {order_days} дн. − конечный остаток {end_balance:g} = {raw_order:.2f}.\n"
                    f"Рекомендуемый заказ: {recommended_order}."
                )
            else:
                if days_left is not None and days_left < order_days:
                    decision_note = "Проверить: риск обнуления до следующего завоза."
                else:
                    decision_note = "Не берём: текущего остатка хватает до следующего завоза."
                comment_text = (
                    f"{base_note}\n"
                    f"{decision_note}\n"
                    f"Расчёт: темп продаж {daily_sales:.2f} × {order_days} дн. − конечный остаток {end_balance:g} = {raw_order:.2f}.\n"
                    "Рекомендуемый заказ: 0."
                )
            recommended_order_cell.comment = Comment(comment_text, "Обнуление")
        except Exception:
            pass

        # Подсветка №1: конечный остаток = 0.
        if end_balance == 0:
            end_cell.fill = copy.copy(RED_FILL)
            end_cell.font = copy.copy(RED_FONT)
            zero_end_count += 1

        # Подсветка №2: остаток товара в днях < 7.
        if days_left is not None and days_left < 7:
            days_left_cell.fill = copy.copy(RED_FILL)
            days_left_cell.font = copy.copy(RED_FONT)
            less_7_count += 1

    first_row = min(rows)
    last_row = max(rows)

    # Условное форматирование добавляем тоже: если потом вручную поменять значения, Excel сам перекрасит.
    ws.conditional_formatting.add(
        f"F{first_row}:F{last_row}",
        FormulaRule(formula=[f'AND(F{first_row}=0,F{first_row}<>"")'], stopIfTrue=False, fill=RED_FILL, font=RED_FONT),
    )
    ws.conditional_formatting.add(
        f"H{first_row}:H{last_row}",
        FormulaRule(formula=[f'AND(H{first_row}<7,H{first_row}<>"")'], stopIfTrue=False, fill=RED_FILL, font=RED_FONT),
    )

    return zero_end_count, less_7_count, order_count, stats



def sort_data_rows_by_days_left(ws, rows):
    """
    Сортирует строки товаров физически в самом листе:
    сначала товары с самым маленьким «Остатком товара в днях»,
    потом товары без продаж/без расчёта. Это надёжнее, чем просто указать
    Excel сортировку в автофильтре: при открытии файл уже выглядит отсортированным.
    """
    if not rows:
        return

    max_col = max(ws.max_column, 9)

    def snapshot_row(row_num):
        cells = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cells.append({
                "value": cell.value,
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
            })
        return {
            "height": ws.row_dimensions[row_num].height,
            "hidden": ws.row_dimensions[row_num].hidden,
            "outlineLevel": ws.row_dimensions[row_num].outlineLevel,
            "cells": cells,
        }

    snapshots = [snapshot_row(row) for row in rows]

    def sort_key(item):
        cells = item["cells"]
        name = str(cells[1]["value"] or "").strip().lower()
        end_balance = to_number(cells[5]["value"])
        days_value = cells[7]["value"]
        days = to_number(days_value) if days_value not in (None, "") else None

        # Пустые значения / нет продаж — вниз.
        no_days = days is None or days_value == ""
        # При равном сроке товары с нулевым остатком выше.
        zero_rank = 0 if end_balance == 0 else 1
        return (1 if no_days else 0, days if days is not None else 10**12, zero_rank, name)

    snapshots.sort(key=sort_key)

    for target_row, item in zip(rows, snapshots):
        ws.row_dimensions[target_row].height = item["height"]
        ws.row_dimensions[target_row].hidden = item["hidden"]
        ws.row_dimensions[target_row].outlineLevel = item["outlineLevel"]
        for col, data in enumerate(item["cells"], start=1):
            cell = ws.cell(row=target_row, column=col)
            cell.value = data["value"]
            cell.font = data["font"]
            cell.fill = data["fill"]
            cell.border = data["border"]
            cell.alignment = data["alignment"]
            cell.number_format = data["number_format"]
            cell.protection = data["protection"]

    # Обновим служебную строку 1С, чтобы было понятно, почему товары идут не по алфавиту.
    for row in range(1, min(ws.max_row, 20) + 1):
        value = ws.cell(row=row, column=2).value
        if value and str(value).strip().lower().startswith("сортировка:"):
            ws.cell(row=row, column=2).value = "Сортировка: Остаток товара в днях (по возрастанию); Номенклатура (по возрастанию);"
            break


def set_autofilter_sort(ws, header_row, rows):
    """Ставит фильтр на таблицу и указывает Excel сортировку по H по возрастанию."""
    first_data_row = min(rows)
    last_data_row = max(rows)
    ws.auto_filter.ref = f"B{header_row + 1}:I{last_data_row}"
    try:
        ws.auto_filter.add_sort_condition(f"H{first_data_row}:H{last_data_row}", descending=False)
    except Exception:
        # Даже если конкретная версия Excel/openpyxl не сохранит sortState,
        # строки уже физически отсортированы функцией выше.
        pass





def refresh_recommended_order_comments_after_sort(ws, rows, fallback_order_days, product_supplier_map=None, supplier_order_info=None):
    """
    Пересоздаёт примечания к «Рекомендуемому заказу» ПОСЛЕ физической сортировки строк.

    Важно: если примечания создавать до сортировки, Excel-комментарии могут остаться на старых
    координатах ячеек и визуально попасть к другому товару. Поэтому финальные комментарии
    формируются заново уже на отсортированной таблице.
    """
    product_supplier_map = product_supplier_map or {}
    supplier_order_info = supplier_order_info or {}

    for row in rows:
        product_name = ws.cell(row=row, column=2).value
        product_norm = norm_text(product_name)
        expense = to_number(ws.cell(row=row, column=5).value)       # Расход, E
        end_balance = to_number(ws.cell(row=row, column=6).value)   # Конечный остаток, F
        daily_sales = to_number(ws.cell(row=row, column=7).value)   # Темп продаж, G
        days_left_value = ws.cell(row=row, column=8).value           # Остаток товара в днях, H
        days_left = to_number(days_left_value) if days_left_value not in (None, "") else None
        recommended_order = to_number(ws.cell(row=row, column=9).value)  # Рекомендуемый заказ, I
        recommended_order_cell = ws.cell(row=row, column=9)

        order_days = fallback_order_days
        supplier_note = ""
        used_fallback_reason = ""

        if product_supplier_map and supplier_order_info:
            supplier = product_supplier_map.get(product_norm)
            if supplier:
                supplier_norm = norm_text(supplier)
                info = supplier_order_info.get(supplier_norm)
                if info:
                    order_days = info["order_days"]
                    delivery_word = "раз" if info["delivery_count"] == 1 else "раза" if 2 <= info["delivery_count"] <= 4 else "раз"
                    supplier_note = (
                        f"Поставщик: {supplier}. Завоз: {info['delivery_count']} {delivery_word} в неделю "
                        f"({info['delivery_days_text']}). Самый длинный промежуток: {info['gap']} дн."
                    )
                    if info.get("safety_days", 0):
                        supplier_note += f" + страховой запас {info['safety_days']} дн."
                else:
                    used_fallback_reason = f"Для поставщика {supplier} не найден график поставок."
            else:
                used_fallback_reason = "Товар не найден в справочнике товар-поставщик."
        elif product_supplier_map or supplier_order_info:
            used_fallback_reason = "Не найден один из общих файлов для расчёта по поставщикам."

        base_note = supplier_note if supplier_note else f"Расчёт по запасному значению: {order_days} дн."
        if used_fallback_reason:
            base_note += f" {used_fallback_reason}"

        if expense <= 0 or daily_sales <= 0:
            if end_balance > 0:
                decision_note = "Не берём: за период не было продаж."
            elif end_balance == 0:
                decision_note = "Не берём: конечный остаток = 0, но за период не было продаж."
            else:
                decision_note = "Проверить: конечный остаток отрицательный, продаж за период не было."
            comment_text = (
                f"{base_note}\n"
                f"{decision_note}\n"
                "Рекомендуемый заказ: 0, потому что темп продаж = 0."
            )
        else:
            raw_order = daily_sales * order_days - end_balance
            if recommended_order > 0:
                if end_balance <= 0:
                    decision_note = "Берём: продажи были, конечный остаток 0/минус. Расчёт до следующего завоза."
                else:
                    decision_note = "Берём: конечного остатка не хватит до следующего завоза."
                comment_text = (
                    f"{base_note}\n"
                    f"{decision_note}\n"
                    f"Расчёт: темп продаж {daily_sales:.2f} × {order_days} дн. − конечный остаток {end_balance:g} = {raw_order:.2f}.\n"
                    f"Рекомендуемый заказ: {int(recommended_order)}."
                )
            else:
                if days_left is not None and days_left < order_days:
                    decision_note = "Проверить: риск обнуления до следующего завоза."
                else:
                    decision_note = "Не берём: конечного остатка хватает до следующего завоза."
                comment_text = (
                    f"{base_note}\n"
                    f"{decision_note}\n"
                    f"Расчёт: темп продаж {daily_sales:.2f} × {order_days} дн. − конечный остаток {end_balance:g} = {raw_order:.2f}.\n"
                    "Рекомендуемый заказ: 0."
                )

        try:
            recommended_order_cell.comment = None
            recommended_order_cell.comment = Comment(comment_text, "Обнуление")
        except Exception:
            pass

def prepare_view_like_example(ws, header_row, rows):
    """
    Делает вид как в готовом примере:
    - служебные строки 1–8 скрыты;
    - сверху остаются заголовки таблицы 9–10;
    - фильтр стоит на строке 10;
    - при открытии файл уже отсортирован по H.
    """
    # В примерах из 1С строки 1–8 служебные. Они мешают работе с фильтром, поэтому скрываем их.
    # Если вдруг заголовок найден не на 9 строке, скрываем всё до верхнего заголовка таблицы.
    last_service_row = max(1, header_row - 1)
    for row in range(1, last_service_row + 1):
        ws.row_dimensions[row].hidden = True

    # Заголовки таблицы должны быть видимыми.
    ws.row_dimensions[header_row].hidden = False
    ws.row_dimensions[header_row + 1].hidden = False

    # Фиксируем окно так, чтобы шапка таблицы оставалась сверху.
    ws.freeze_panes = f"B{header_row + 2}"

    # Ставим фильтр именно на строку подзаголовков, где есть Начальный остаток/Приход/Расход/Конечный остаток/Темп продаж/Остаток товара в днях.
    set_autofilter_sort(ws, header_row, rows)


def process_file(input_path, output_folder, order_days, product_supplier_map=None, supplier_order_info=None):
    wb = load_workbook_safely(input_path)
    ws = wb.active

    period_days = parse_period_days(ws)
    header_row = find_header_row(ws)
    ensure_extra_columns(ws, header_row)
    rows = find_data_rows(ws, header_row)

    clear_existing_conditional_formatting(ws)
    clear_old_highlighting(ws, rows)
    zero_end_count, less_7_count, order_count, match_stats = calculate_and_highlight(
        ws,
        rows,
        period_days,
        order_days,
        product_supplier_map=product_supplier_map,
        supplier_order_info=supplier_order_info,
    )

    # Сортируем товары сразу в готовом файле: критичные по дням будут сверху.
    sort_data_rows_by_days_left(ws, rows)

    # Примечания к заказу создаём ПОСЛЕ сортировки, чтобы комментарии не съезжали на другие товары.
    refresh_recommended_order_comments_after_sort(
        ws,
        rows,
        order_days,
        product_supplier_map=product_supplier_map,
        supplier_order_info=supplier_order_info,
    )

    # Вид и фильтр как в готовом примере.
    prepare_view_like_example(ws, header_row, rows)

    base_name = os.path.basename(input_path)
    name, ext = os.path.splitext(base_name)
    output_name = f"готово_{name}{ext}"
    output_path = os.path.join(output_folder, output_name)

    # Если такой файл уже есть, не перезаписываем молча.
    if os.path.exists(output_path):
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_folder, f"готово_{name}_{stamp}{ext}")

    wb.save(output_path)
    return {
        "file": base_name,
        "output": output_path,
        "rows": len(rows),
        "period_days": period_days,
        "zero_end": zero_end_count,
        "less_7": less_7_count,
        "order_count": order_count,
        "order_days": order_days,
        "match_stats": match_stats,
    }


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Обнуление товаров — анализ запасов")
        self.root.geometry("780x520")
        self.files = []
        self.output_folder = ""
        self.supplier_file = default_file_path(DEFAULT_SUPPLIER_FILENAMES)
        self.schedule_file = default_file_path(DEFAULT_SCHEDULE_FILENAMES)

        tk.Label(root, text="Обнуление товаров", font=("Arial", 16, "bold")).pack(pady=10)

        tk.Label(
            root,
            text="Программа добавит «Темп продаж», «Остаток товара в днях» и «Рекомендуемый заказ». Справочник supplier_map.xlsx и график delivery_schedule.xlsx берутся автоматически из папки программы. К рекомендуемому заказу добавляется примечание с логикой: берём / не берём / проверить. Примечания ставятся после сортировки.",
            wraplength=680,
            justify="center",
        ).pack(pady=4)

        buttons_frame = tk.Frame(root)
        buttons_frame.pack(pady=10)

        tk.Button(buttons_frame, text="1. Выбрать Excel-файлы", width=28, command=self.choose_files).grid(row=0, column=0, padx=8, pady=4)
        tk.Button(buttons_frame, text="2. Выбрать папку", width=28, command=self.choose_folder).grid(row=0, column=1, padx=8, pady=4)

        self.files_label = tk.Label(root, text="Файлы обнуления не выбраны", wraplength=680, justify="left")
        self.files_label.pack(pady=5)

        self.folder_label = tk.Label(root, text="Папка не выбрана", wraplength=680, justify="left")
        self.folder_label.pack(pady=5)

        defaults_status = []
        defaults_status.append("supplier_map.xlsx найден" if self.supplier_file else "supplier_map.xlsx не найден — будет запасное значение дней")
        defaults_status.append("delivery_schedule.xlsx найден" if self.schedule_file else "delivery_schedule.xlsx не найден — будет запасное значение дней")
        self.defaults_label = tk.Label(root, text="; ".join(defaults_status), wraplength=680, justify="left")
        self.defaults_label.pack(pady=5)

        order_frame = tk.Frame(root)
        order_frame.pack(pady=6)
        tk.Label(order_frame, text="Запасное значение, если товар/поставщик не найден, дней:").grid(row=0, column=0, padx=6, sticky="e")
        self.order_days_var = tk.StringVar(value="7")
        tk.Entry(order_frame, textvariable=self.order_days_var, width=8, justify="center").grid(row=0, column=1, padx=6)

        tk.Label(order_frame, text="Страховой запас к графику, дней:").grid(row=1, column=0, padx=6, sticky="e")
        self.safety_days_var = tk.StringVar(value="0")
        tk.Entry(order_frame, textvariable=self.safety_days_var, width=8, justify="center").grid(row=1, column=1, padx=6)

        self.progress_label = tk.Label(root, text="", wraplength=680, justify="center")
        self.progress_label.pack(pady=4)

        self.run_button = tk.Button(
            root,
            text="3. Сформировать готовые файлы",
            width=34,
            height=2,
            bg="#B00000",
            fg="white",
            command=self.run,
        )
        self.run_button.pack(pady=12)

    def choose_files(self):
        files = filedialog.askopenfilenames(
            title="Выбери Excel-файлы из 1С",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if files:
            self.files = list(files)
            names = "\n".join(os.path.basename(path) for path in self.files[:6])
            if len(self.files) > 6:
                names += f"\n...и ещё {len(self.files) - 6}"
            self.files_label.config(text=f"Файлы обнуления: {len(self.files)}\n{names}")

    def choose_folder(self):
        folder = filedialog.askdirectory(title="Выбери папку для готовых файлов")
        if folder:
            self.output_folder = folder
            self.folder_label.config(text=f"Папка: {folder}")

    def run(self):
        if not self.files:
            messagebox.showwarning("Нет файлов", "Сначала выбери Excel-файлы обнуления.")
            return
        if not self.output_folder:
            messagebox.showwarning("Нет папки", "Сначала выбери папку для сохранения.")
            return

        try:
            order_days = int(str(self.order_days_var.get()).strip())
            if order_days <= 0:
                raise ValueError
            self.order_days = order_days
        except Exception:
            messagebox.showwarning("Неверные дни", "В поле запасного значения нужно указать целое число больше 0. Например: 7.")
            return

        try:
            safety_days = int(str(self.safety_days_var.get()).strip())
            if safety_days < 0:
                raise ValueError
            self.safety_days = safety_days
        except Exception:
            messagebox.showwarning("Неверный страховой запас", "В поле страхового запаса нужно указать целое число 0 или больше. Например: 0, 1, 2.")
            return

        self.run_button.config(state="disabled", text="Идёт обработка...")
        self.progress_label.config(text="Начинаю обработку. Окно можно не трогать — файлы считаются в фоне.")

        worker = threading.Thread(target=self._run_worker, daemon=True)
        worker.start()

    def _set_progress(self, text):
        self.root.after(0, lambda: self.progress_label.config(text=text))

    def _finish(self, text):
        def show_result():
            self.run_button.config(state="normal", text="3. Сформировать готовые файлы")
            self.progress_label.config(text="Обработка завершена.")
            messagebox.showinfo("Результат", text)
        self.root.after(0, show_result)

    def _run_worker(self):
        results = []
        errors = []
        total = len(self.files)

        product_supplier_map = {}
        supplier_order_info = {}
        supplier_map_count = 0
        schedule_count = 0

        # Готовим общие файлы один раз, а не для каждого отчёта.
        if self.supplier_file and self.schedule_file:
            try:
                self._set_progress("Читаю общий справочник товар-поставщик...")
                product_supplier_map = parse_supplier_product_map(self.supplier_file)
                supplier_map_count = len(product_supplier_map)

                self._set_progress("Читаю график поставок...")
                schedule = parse_delivery_schedule(self.schedule_file)
                supplier_order_info = build_supplier_order_info(schedule, self.safety_days)
                schedule_count = len(supplier_order_info)
            except Exception as error:
                errors.append(f"Общие файлы: {error}")
                product_supplier_map = {}
                supplier_order_info = {}

        elif self.supplier_file or self.schedule_file:
            errors.append("Для расчёта по поставщикам нужно выбрать оба файла: и справочник товар-поставщик, и график поставок. Сейчас заказ посчитан по запасному значению дней.")

        for index, path in enumerate(self.files, start=1):
            name = os.path.basename(path)
            self._set_progress(f"Обрабатываю {index} из {total}: {name}")
            try:
                results.append(process_file(
                    path,
                    self.output_folder,
                    self.order_days,
                    product_supplier_map=product_supplier_map,
                    supplier_order_info=supplier_order_info,
                ))
            except Exception as error:
                errors.append(f"{name}: {error}")

        if results:
            total_rows = sum(item["rows"] for item in results)
            total_zero = sum(item["zero_end"] for item in results)
            total_less_7 = sum(item["less_7"] for item in results)
            total_order = sum(item["order_count"] for item in results)
            stats = {
                "supplier_found": 0,
                "schedule_found": 0,
                "fallback_used": 0,
                "product_not_found": 0,
                "supplier_without_schedule": 0,
            }
            for item in results:
                for key, value in item.get("match_stats", {}).items():
                    stats[key] = stats.get(key, 0) + value

            if product_supplier_map and supplier_order_info:
                order_mode = (
                    f"по графику поставщиков\n"
                    f"Товаров в справочнике: {supplier_map_count}\n"
                    f"Поставщиков в графике: {schedule_count}\n"
                    f"Товаров, где найден поставщик: {stats['supplier_found']}\n"
                    f"Товаров, где найден график поставщика: {stats['schedule_found']}\n"
                    f"Товаров, посчитанных по запасному значению {self.order_days} дней: {stats['fallback_used']}"
                )
            else:
                order_mode = f"по общему запасному значению: {self.order_days} дней"

            text = (
                f"Готово.\n\n"
                f"Обработано файлов: {len(results)}\n"
                f"Товаров: {total_rows}\n"
                f"Конечный остаток = 0: {total_zero}\n"
                f"Остаток товара в днях < 7: {total_less_7}\n"
                f"Рекомендуемый заказ > 0: {total_order}\n"
                f"Расчёт рекомендуемого заказа: {order_mode}\n\n"
                f"Файлы сохранены в выбранную папку."
            )
        else:
            text = "Не получилось обработать файлы."

        if errors:
            text += "\n\nПредупреждения/ошибки:\n" + "\n".join(errors[:8])
            if len(errors) > 8:
                text += f"\n...и ещё {len(errors) - 8}"

        self._finish(text)


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
