from datetime import date
from pathlib import Path
import importlib.util

import pytest
from openpyxl import Workbook, load_workbook


APP_PATH = Path(__file__).resolve().parents[1] / "programma_obnulenie_app.py"
spec = importlib.util.spec_from_file_location("obnulenie_app", APP_PATH)
app = importlib.util.module_from_spec(spec)
spec.loader.exec_module(app)


@pytest.mark.parametrize(
    ("text", "expected_days"),
    [
        ("Период: 01.07.2026 - 07.07.2026", 7),
        ("Период: 16 июля 2026 г.", 1),
        ("Период: 12–15 июля 2026 г.", 4),
        ("Период: 31.12.2025 - 02.01.2026", 3),
    ],
)
def test_period_is_inclusive(text, expected_days):
    dates = app._dates_from_text(text)
    assert dates
    assert app._period_length(dates[0], dates[-1]) == expected_days


def test_max_gap_between_delivery_days():
    assert app.max_gap_between_delivery_days({0}) == 7
    assert app.max_gap_between_delivery_days({0, 3}) == 4
    assert app.max_gap_between_delivery_days({0, 1, 2, 3, 4, 5}) == 2


def test_process_file_calculates_daily_sales_and_order(tmp_path):
    input_path = tmp_path / "report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["B2"] = "Период: 01.07.2026 - 07.07.2026"
    ws["B9"] = "Номенклатура"
    ws.merge_cells("C9:F9")
    ws["C9"] = "Количество (в базовых единицах)"
    for column, value in enumerate(
        ["Номенклатура", "Начальный остаток", "Приход", "Расход", "Конечный остаток"],
        start=2,
    ):
        ws.cell(10, column, value)

    ws.cell(11, 2, "Тестовый товар")
    ws.cell(11, 3, 10)
    ws.cell(11, 4, 6)
    ws.cell(11, 5, 14)
    ws.cell(11, 6, 2)
    ws.cell(12, 2, "Итог")
    wb.save(input_path)

    result = app.process_file(str(input_path), str(tmp_path), 7)
    output = load_workbook(result["output"])
    out_ws = output.active

    assert result["period_days"] == 7
    assert out_ws["G11"].value == 2
    assert out_ws["H11"].value == 1
    assert out_ws["I11"].value == 12
    assert out_ws["I11"].comment is not None


def test_build_supplier_order_info_with_safety_days():
    info = app.build_supplier_order_info({"поставщик": {0, 3}}, safety_days=1)
    assert info["поставщик"]["gap"] == 4
    assert info["поставщик"]["order_days"] == 5
